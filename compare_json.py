# compare_json.py

import json
import logging
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from utils import adjust_column_widths

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)


def load_json(file_path: str) -> Any:
    """
    Загружает JSON данные из файла.

    :param file_path: Путь к JSON файлу.
    :return: Загруженные данные.
    :raises FileNotFoundError: Если файл не найден.
    :raises json.JSONDecodeError: Если файл содержит некорректный JSON.
    :raises Exception: Для остальных ошибок.
    """
    try:
        logging.info(f"Загружаю данные из JSON-файла: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except json.JSONDecodeError as jde:
        logging.error(f"Ошибка декодирования JSON файла {file_path}: {jde}")
        raise
    except FileNotFoundError as fnfe:
        logging.error(f"Файл не найден {file_path}: {fnfe}")
        raise
    except Exception as e:
        logging.error(f"Не удалось загрузить JSON файл {file_path}: {e}")
        raise


def build_dict1(data1: Any) -> Dict[str, Dict[str, Any]]:
    """
    Преобразует данные из первого JSON-файла (паспорт) в словарь.

    :param data1: Данные из первого JSON-файла.
    :return: Словарь с именами серверов в нижнем регистре в качестве ключей.
    """
    dict1: Dict[str, Dict[str, Any]] = {}
    for section in data1:
        section_name = section.get('Раздел', '')
        logging.debug(f"Раздел: {section_name}")
        for item in section.get('Данные', []):
            for vm in item.get('ВМ', []):
                server_name = vm.get('Имя сервера', '').strip().lower()
                if server_name:
                    dict1[server_name] = {
                        'IP адрес': vm.get('IP адрес', ''),
                        'Сайзинг': vm.get('Сайзинг', ''),
                        'Источник': f"Раздел: {section_name}"
                    }
    return dict1


def build_dict2(data2: Any) -> Dict[str, Dict[str, Any]]:
    """
    Преобразует данные из второго JSON-файла (сайзинг) в словарь.

    :param data2: Данные из второго JSON-файла.
    :return: Словарь с именами серверов в нижнем регистре в качестве ключей.
    """
    return {
        item.get('Имя сервера', '').strip().lower(): {
            'IP адрес': item.get('IP адрес', ''),
            'Сайзинг': item.get('Сайзинг', ''),
            'Источник': "Excel файл"
        }
        for item in data2
        if item.get('Имя сервера', '').strip()
    }


def compare_json(json_file_1: str, json_file_2: str, output_excel_file: str) -> None:
    """
    Сравнивает два JSON файла и записывает результаты сравнения в Excel файл.

    :param json_file_1: Путь к первому JSON файлу (паспорт).
    :param json_file_2: Путь ко второму JSON файлу (сайзинг).
    :param output_excel_file: Путь к выходному Excel файлу.
    """
    try:
        # Загружаем данные из JSON-файлов
        data1 = load_json(json_file_1)
        data2 = load_json(json_file_2)

        # Преобразуем данные в словари для быстрого доступа
        dict1 = build_dict1(data1)
        dict2 = build_dict2(data2)

        # Получаем множество всех серверов
        servers1 = set(dict1.keys())
        servers2 = set(dict2.keys())
        all_servers = servers1.union(servers2)
        logging.info(f"Всего серверов в паспорте: {len(servers1)}")
        logging.info(f"Всего серверов в сайзинге: {len(servers2)}")
        logging.info(f"Общее количество уникальных серверов для сравнения: {len(all_servers)}")

        # Подготавливаем данные для записи в Excel
        matched_rows: List[Dict[str, Any]] = []
        unmatched_rows_red: List[Dict[str, Any]] = []
        unmatched_rows_blue: List[Dict[str, Any]] = []

        for server in sorted(all_servers):
            logging.info(f"\nСравнение сервера: {server}")
            row: Dict[str, Any] = {'Имя сервера': server}
            red_cells: List[str] = []
            blue_cells: List[str] = []
            full_row_color: Optional[str] = None

            item1 = dict1.get(server)
            item2 = dict2.get(server)

            # Добавляем данные из паспорта
            if item1:
                row['IP адрес в паспорте'] = item1.get('IP адрес', '')
                row['Сайзинг в паспорте'] = item1.get('Сайзинг', '')
                row['Источник в паспорте'] = item1.get('Источник', '')
                logging.debug(
                    f"  Данные из паспорта: IP адрес: {row['IP адрес в паспорте']}, "
                    f"Сайзинг: {row['Сайзинг в паспорте']}, Источник: {row['Источник в паспорте']}"
                )
            else:
                row['IP адрес в паспорте'] = ''
                row['Сайзинг в паспорте'] = ''
                row['Источник в паспорте'] = ''
                logging.debug("  Сервер отсутствует в паспорте")

            # Добавляем данные из сайзинга
            if item2:
                row['IP адрес в сайзинге'] = item2.get('IP адрес', '')
                row['Сайзинг в сайзинге'] = item2.get('Сайзинг', '')
                logging.debug(
                    f"  Данные из сайзинга: IP адрес: {row['IP адрес в сайзинге']}, "
                    f"Сайзинг: {row['Сайзинг в сайзинге']}, Источник: {item2.get('Источник', '')}"
                )
            else:
                row['IP адрес в сайзинге'] = ''
                row['Сайзинг в сайзинге'] = ''
                logging.debug("  Сервер отсутствует в сайзинге")

            # Логика подсветки
            if item1 and item2:
                discrepancies = False
                ip1 = str(row['IP адрес в паспорте']).strip()
                ip2 = str(row['IP адрес в сайзинге']).strip()
                if ip1 != ip2:
                    discrepancies = True
                    red_cells.append('IP адрес в паспорте')

                sizing1 = str(row['Сайзинг в паспорте']).strip()
                sizing2 = str(row['Сайзинг в сайзинге']).strip()
                if sizing1 != sizing2:
                    discrepancies = True
                    red_cells.append('Сайзинг в паспорте')

                matched_rows.append({
                    'data': row,
                    'red_cells': red_cells,
                    'blue_cells': blue_cells,
                    'full_row_color': None
                })
            else:
                if not item1:
                    full_row_color = 'red'
                    logging.info("  Сервер отсутствует в паспорте")
                if not item2:
                    full_row_color = 'blue'
                    logging.info("  Сервер отсутствует в сайзинге")

                if not item1 and not item2:
                    full_row_color = 'red'  # При отсутствии в обоих, выделяем красным

                row_entry = {
                    'data': row,
                    'red_cells': red_cells,
                    'blue_cells': blue_cells,
                    'full_row_color': full_row_color
                }

                if full_row_color == 'red':
                    unmatched_rows_red.append(row_entry)
                elif full_row_color == 'blue':
                    unmatched_rows_blue.append(row_entry)

        # Объединяем списки: сначала совпадающие, затем отсутствующие в паспорте (красным), затем отсутствующие в сайзинге (синим)
        all_rows = matched_rows + unmatched_rows_red + unmatched_rows_blue

        # Записываем результаты в Excel-файл
        wb = Workbook()
        ws = wb.active

        headers = [
            'Имя сервера',
            'IP адрес в сайзинге', 'IP адрес в паспорте',
            'Сайзинг в сайзинге', 'Сайзинг в паспорте',
            'Источник в паспорте'
        ]

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Светло-синий
        bold_font = Font(bold=True)

        current_row = 1

        # Пишем совпадающие серверы
        ws.cell(row=current_row, column=1, value="Совпадающие серверы").font = bold_font
        current_row += 1
        ws.append(headers)
        for entry in matched_rows:
            row_data = entry['data']
            ws_row = [row_data.get(header, '') for header in headers]
            ws.append(ws_row)
            row_num = ws.max_row
            for col_name in entry['red_cells']:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    ws.cell(row=row_num, column=col_idx).fill = red_fill
        current_row = ws.max_row + 2

        # Пишем серверы отсутствующие в паспорте
        ws.cell(row=current_row, column=1, value="Серверы отсутствующие в паспорте").font = bold_font
        current_row += 1
        ws.append(headers)
        for entry in unmatched_rows_red:
            row_data = entry['data']
            ws_row = [row_data.get(header, '') for header in headers]
            ws.append(ws_row)
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = red_fill
        current_row = ws.max_row + 2

        # Пишем серверы отсутствующие в сайзинге
        ws.cell(row=current_row, column=1, value="Серверы отсутствующие в сайзинге").font = bold_font
        current_row += 1
        ws.append(headers)
        for entry in unmatched_rows_blue:
            row_data = entry['data']
            ws_row = [row_data.get(header, '') for header in headers]
            ws.append(ws_row)
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = blue_fill

        # Настраиваем ширину колонок
        adjust_column_widths(ws)

        # Сохраняем Excel-файл
        wb.save(output_excel_file)
        logging.info(f"\nРезультаты сравнения сохранены в файле {output_excel_file}")

    except json.JSONDecodeError as jde:
        logging.error(f"Ошибка декодирования JSON: {jde}")
        raise
    except FileNotFoundError as fnfe:
        logging.error(f"Файл не найден: {fnfe}")
        raise
    except Exception as e:
        logging.error(f"Неизвестная ошибка при сравнении JSON-файлов: {e}")
        raise
