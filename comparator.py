# comparator.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging
from utils import adjust_column_widths


def compare_data(df_json, df_excel, output_excel_file):
    try:
        # Приводим имена серверов к единому регистру для корректного сравнения
        df_json['Имя сервера'] = df_json['Имя сервера'].str.strip().str.lower()
        df_excel['Имя сервера'] = df_excel['Имя сервера'].str.strip().str.lower()

        # Сравниваем по имени сервера
        merged_df = df_excel.merge(df_json, on='Имя сервера', how='left', suffixes=('_excel', '_json'), indicator=True)

        # Убедимся, что нужные колонки существуют
        for col in ['Сайзинг_json', 'IP адрес_json']:
            if col not in merged_df.columns:
                merged_df[col] = None

        # Инициализируем словарь для отслеживания ячеек, которые нужно раскрасить
        cells_to_color = {}

        # Проверяем соответствие IP адресов и сайзинга
        for index, row in merged_df.iterrows():
            vm_name = row['Имя сервера']
            ip_excel = row['IP адрес_excel']
            ip_json = row['IP адрес_json']
            sizing_excel = row['Сайзинг_excel']
            sizing_json = row['Сайзинг_json']

            logging.info(f"ВМ {vm_name} -> поиск в паспорте")

            # Если ВМ не найдена в паспорте
            if row['_merge'] == 'left_only' or pd.isna(ip_json):
                logging.info(f"ВМ {vm_name} -> не найдена в паспорте")
                # Красим всю строку в красный
                cells_to_color[index] = {'full_row': True}
                continue
            else:
                logging.info(f"ВМ {vm_name} -> найдена в паспорте")
                # Инициализируем список ячеек для окраски в этой строке
                cells_to_color[index] = {'full_row': False, 'cells': []}

            # Проверяем соответствие IP адреса
            logging.info(f"ВМ {vm_name} -> сверяю IP адрес")
            if pd.notna(ip_excel) and pd.notna(ip_json):
                if ip_excel.strip() == ip_json.strip():
                    logging.info(f"ВМ {vm_name} -> IP адрес {ip_excel.strip()} совпадает")
                else:
                    logging.info(
                        f"ВМ {vm_name} -> IP адрес не совпадает: Excel - {ip_excel.strip()}, Паспорт - {ip_json.strip()}")
                    # Помечаем ячейку IP адреса в колонке IP адрес_json для окраски
                    cells_to_color[index]['cells'].append('IP адрес_json')
            else:
                logging.info(f"ВМ {vm_name} -> IP адрес отсутствует в одном из источников")

            # Проверяем сайзинг
            logging.info(f"ВМ {vm_name} -> сверяю сайзинг")
            if pd.notna(sizing_excel) and pd.notna(sizing_json):
                if sizing_excel.strip() == sizing_json.strip():
                    logging.info(f"ВМ {vm_name} -> сайзинг {sizing_excel.strip()} совпадает")
                else:
                    logging.info(
                        f"ВМ {vm_name} -> сайзинг не совпадает: Excel - {sizing_excel.strip()}, Паспорт - {sizing_json.strip()}")
                    # Помечаем ячейку Сайзинг в колонке Сайзинг_json для окраски
                    cells_to_color[index]['cells'].append('Сайзинг_json')
            else:
                logging.info(f"ВМ {vm_name} -> сайзинг отсутствует в одном из источников")

        # После обработки, удаляем колонку '_merge'
        merged_df.drop(columns=['_merge'], inplace=True)

        # Подготавливаем колонки в нужном порядке
        columns_order = ['Имя сервера', 'Сайзинг_excel', 'IP адрес_excel', 'IP адрес_json', 'Сайзинг_json']
        merged_df = merged_df[columns_order + [col for col in merged_df.columns if col not in columns_order]]

        # Теперь записываем DataFrame в Excel с учетом форматирования
        wb = Workbook()
        ws = wb.active

        # Записываем заголовки
        headers = list(merged_df.columns)
        ws.append(headers)

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Записываем данные и применяем форматирование
        for idx, row in merged_df.iterrows():
            row_values = []
            for col in headers:
                row_values.append(row.get(col, ''))
            ws.append(row_values)
            row_num = ws.max_row

            # Применяем цветовое выделение
            if cells_to_color.get(idx, {}).get('full_row'):
                # Окрашиваем всю строку
                for cell in ws[row_num]:
                    cell.fill = red_fill
            else:
                # Окрашиваем конкретные ячейки
                for col_name in cells_to_color.get(idx, {}).get('cells', []):
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1  # +1, так как индексация openpyxl начинается с 1
                        ws.cell(row=row_num, column=col_idx).fill = red_fill

        # Настраиваем ширину колонок
        adjust_column_widths(ws)

        # Сохраняем файл
        wb.save(output_excel_file)
        logging.info(f"Результаты сравнения сохранены в файле {output_excel_file}")

    except Exception as e:
        logging.error(f"Ошибка при сравнении данных: {e}")
        raise e  # Пробрасываем исключение вверх
